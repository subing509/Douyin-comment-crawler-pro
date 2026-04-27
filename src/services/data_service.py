# -*- coding: utf-8 -*-
"""
数据服务 - 数据导出和管理
"""

import os
import glob
from typing import List, Dict, Optional
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment


class DataService:
    """数据管理服务"""

    COLUMN_MAPPING = {
        'comment_id': '评论ID',
        'video_id': '视频ID',
        'content': '评论内容',
        'user_nickname': '用户昵称',
        'user_id': '用户ID',
        'like_count': '点赞数',
        'reply_count': '回复数',
        'create_time': '发布时间',
        'ip_location': 'IP属地',
        'is_author': '是否作者',
        'video_url': '视频链接',
    }

    PREFERRED_COLUMNS = [
        '评论内容', '用户昵称', '点赞数', '回复数',
        '发布时间', 'IP属地', '是否作者',
        '视频标题', '视频作者', '视频链接', '视频ID', '评论ID'
    ]
    
    def __init__(self, output_dir: str = "outputs"):
        """初始化数据服务
        
        Args:
            output_dir: 输出目录路径
        """
        self.output_dir = output_dir
        os.makedirs(self.output_dir, exist_ok=True)
    
    def export_to_excel(
        self,
        comments: List[Dict],
        video_info: Dict,
        filename: str
    ):
        """导出数据到Excel
        
        Args:
            comments: 评论数据列表
            video_info: 视频信息
            filename: 输出文件名
        """
        if not comments:
            print("⚠️ 没有评论数据可导出")
            return
        
        print(f"💾 正在导出数据到 {filename}...")
        
        # 创建DataFrame
        df = pd.DataFrame(comments)
        df = self._ensure_internal_columns(df)
        
        # 添加视频信息列
        df['视频标题'] = video_info.get('title', '')
        df['视频作者'] = video_info.get('author_nickname', '')
        df['视频链接'] = video_info.get('video_url', '')
        
        # 重命名列
        df = df.rename(columns=self.COLUMN_MAPPING)
        
        # 调整列顺序
        existing_columns = [col for col in self.PREFERRED_COLUMNS if col in df.columns]
        df = df[existing_columns]
        
        # 导出到Excel
        filepath = os.path.join(self.output_dir, filename)
        df.to_excel(filepath, index=False, engine='openpyxl')
        
        # 格式化Excel
        self._format_excel(filepath)
        
        print(f"✅ 成功导出 {len(comments)} 条评论到 {filepath}")
    
    def _format_excel(self, filepath: str):
        """格式化Excel文件
        
        Args:
            filepath: Excel文件路径
        """
        try:
            wb = load_workbook(filepath)
            ws = wb.active
            
            # 设置表头样式
            header_font = Font(bold=True, size=11)
            header_alignment = Alignment(horizontal='center', vertical='center')
            
            for cell in ws[1]:
                cell.font = header_font
                cell.alignment = header_alignment
            
            # 自动调整列宽
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                
                for cell in column:
                    try:
                        if cell.value:
                            cell_length = len(str(cell.value))
                            if cell_length > max_length:
                                max_length = cell_length
                    except:
                        pass
                
                # 设置列宽（最大60）
                adjusted_width = min(max_length + 2, 60)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            wb.save(filepath)
        except Exception as e:
            print(f"⚠️ 格式化Excel失败: {e}")

    def load_existing_data(self, video_id: str) -> Optional[pd.DataFrame]:
        """加载已有数据
        
        Args:
            video_id: 视频ID
            
        Returns:
            已有数据DataFrame，不存在则返回None
        """
        # 查找包含该视频ID的Excel文件
        pattern = os.path.join(self.output_dir, f"video_{video_id}_*.xlsx")
        files = glob.glob(pattern)
        
        if not files:
            return None
        
        # 使用最新的文件
        latest_file = max(files, key=os.path.getctime)
        
        try:
            df = pd.read_excel(latest_file, engine='openpyxl')
            df = self._to_internal_schema(df)
            print(f"📂 加载历史数据: {latest_file} ({len(df)} 条评论)")
            return df
        except Exception as e:
            print(f"⚠️ 加载历史数据失败: {e}")
            return None
    
    def merge_incremental_data(
        self,
        existing_df: pd.DataFrame,
        new_comments: List[Dict]
    ) -> pd.DataFrame:
        """合并增量数据
        
        Args:
            existing_df: 已有数据
            new_comments: 新采集的评论
            
        Returns:
            合并后的DataFrame
        """
        if not new_comments:
            return existing_df
        
        # 创建新数据DataFrame
        new_df = pd.DataFrame(new_comments)
        
        existing_df = self._ensure_internal_columns(existing_df)
        new_df = self._ensure_internal_columns(new_df)

        merged_df = pd.concat([existing_df, new_df], ignore_index=True)
        before = len(existing_df)
        merged_df = merged_df.drop_duplicates(subset=['comment_id'], keep='first')
        print(f"✅ 合并数据: 原有 {before} 条，新增 {len(merged_df) - before} 条")
        return merged_df
    
    def generate_report(self, results: List[Dict]) -> str:
        """生成采集报告
        
        Args:
            results: 采集结果列表
            
        Returns:
            报告文本
        """
        if not results:
            return "没有采集结果"
        
        total = len(results)
        success = sum(1 for r in results if r.get('status') == 'success')
        failed = sum(1 for r in results if r.get('status') == 'failed')
        total_comments = sum(r.get('comments_count', 0) for r in results)
        new_comments = sum(r.get('new_comments_count', 0) for r in results)
        
        report = f"""
╔══════════════════════════════════════╗
║          采集报告                     ║
╠══════════════════════════════════════╣
║ 总任务数: {total:>4} 个                  ║
║ 成功数:   {success:>4} 个                  ║
║ 失败数:   {failed:>4} 个                  ║
║ 总评论数: {total_comments:>4} 条                  ║
║ 新增评论: {new_comments:>4} 条                  ║
╚══════════════════════════════════════╝
"""
        return report

    # ------------------------------------------------------------------
    # 内部工具
    # ------------------------------------------------------------------

    def _ensure_internal_columns(self, df: pd.DataFrame) -> pd.DataFrame:
        """确保DataFrame使用内部字段命名"""
        if df is None or df.empty:
            return pd.DataFrame()
        return df.rename(columns=self._reverse_column_mapping())

    def _to_internal_schema(self, df: pd.DataFrame) -> pd.DataFrame:
        """把导出的中文列转换回内部字段"""
        if df is None or df.empty:
            return pd.DataFrame()
        reverse = self._reverse_column_mapping()
        normalized = df.rename(columns=reverse)
        return normalized

    def _reverse_column_mapping(self) -> Dict[str, str]:
        """获取中文->英文的映射"""
        return {v: k for k, v in self.COLUMN_MAPPING.items()}
